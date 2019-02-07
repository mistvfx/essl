import pymysql
import pandas as pd
import numpy as np
from pandas import *
import datetime
import calendar
import threading
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.progressbar import ProgressBar
from kivy.core.image import Image
from kivy.uix.button import Button

from db.calcWrkHrs import calActualWorkingHours
from db import *
from pages import *

import openpyxl

from kivy.lang import Builder

Builder.load_string("""
<ExcelLoading>:
    source: 'icons/export.gif'
    size: self.parent.size
    y: self.parent.y
    x: self.parent.x
    keep_data: True
""")

class ExcelLoading(Image):
    pass

def excelUpPB():
    exceLoad = excelLoading()
    popup = Popup(content=exceLoad, size_hint=(0.5, 0.5))
    popup.open()

    return 0

def formatTime(time):
    seconds = time.total_seconds()
    hours = int(seconds / 3600)
    minutes = int((seconds % 3600) / 60)
    seconds = int(seconds % 60)

    return ('{}:{}:{}'.format(hours, minutes, seconds))

def excelManip(filePath):
    db = pymysql.connect("10.10.5.60", "mcheck", "mcheck@123", "essl", autocommit=True, connect_timeout=1)
    cur = db.cursor()
    cur1 = db.cursor()
    print(filePath[0])

    df = pd.read_excel(filePath[0], sheet_name='data')

    for i in df.index:
        id = df['Postcode'][i]
        artist = df['First Name'][i]
        dept = df['City'][i]
        reader_name = df['Reader Name'][i]
        date_time = df['Time'][i]
        door = df['Event Point'][i]
        event = df['Event Description'][i]
        try:
            reader = reader_name.split(" ")
            io = reader[1]
        except AttributeError:
            continue
        date = date_time.date()
        time = date_time.time()
        #print(type(time))  if str(id) != 'nan' and str(event) != 'Anti-Passback':

        if str(id) != 'nan':
            try:
                cur.execute("INSERT INTO essl.%d (IO, MTIME, MDATE, DOOR, AccType) VALUES('%s', '%s', '%s', '%s', '%s')" %(id, io, time, date, door, event))
            except:
                cur.execute("INSERT INTO essl.user_master(ID, Name, Department, Password, Level) VALUES('%d','%s','%s','%d','5');" %(id, artist, dept, id))
                cur.execute("CREATE TABLE IF NOT EXISTS essl.%d (SNum int(11) NOT NULL AUTO_INCREMENT, IO char(4) NOT NULL, MTIME time NOT NULL, MDATE date NOT NULL, DOOR varchar(45) NOT NULL, AccType varchar(50) NOT NULL, PRIMARY KEY (`SNum`))" %(id))
                cur.execute("INSERT INTO essl.%d (IO, MTIME, MDATE, DOOR, AccType) VALUES('%s', '%s', '%s', '%s', '%s')" %(id, io, time, date, door, event))
            try:
                cur1.execute("INSERT INTO essl.`leaves` (ID) VALUES('%d')"%(int(id)))
            except:
                pass

    cur.close()
    db.close()
    print("EXCEL UPLOAD COMPLETE")
    return 0

def threads(filePath):
    t1 = threading.Thread(target=excelManip, args=(filePath,))
    t2 = threading.Thread(target=excelUpPB)

    t2.start()
    t1.start()

def formatDate(date):
    Date = date.split(".")
    dt = (str(Date[2])+"-"+str(Date[1])+"-"+str(Date[0]))
    return dt

def excelExport(date):
    StdWrkHrs = datetime.timedelta(hours=8, minutes=29, seconds=59)
    absent = datetime.timedelta()
    db = pymysql.connect("10.10.5.60", "mcheck", "mcheck@123", "essl", autocommit=True, connect_timeout=1)
    cur = db.cursor()
    cur.execute("SELECT ID, Name, Department, Level FROM essl.user_master WHERE ID != '1000' AND Status = 'OPEN'")
    cur1 = db.cursor()

    Artist_Code = []
    Artist_Name = []
    Department = []
    Level = []

    ios = []
    timings = []
    doors = []

    In_Time = []
    Out_Time = []
    Work_Duration = []
    Total_Duration = []
    Non_Completed_Hours = []
    Additional_Hours = []
    Remarks = []

    for id in cur.fetchall():
        cur1.execute("SELECT IO, MTIME, DOOR FROM essl.`%d` WHERE MDATE = '%s' ORDER BY MTIME ASC" %(id[0], formatDate(date)))
        #ios = []
        #timings = []
        #print('id:', id[0])
        #doors = []
        for data in cur1.fetchall():
            ios.append(data[0])
            timings.append(data[1])
            doors.append(data[2])
        Artist_Code.append(id[0])
        Artist_Name.append(id[1])
        Department.append(id[2])
        Level.append(id[3])
        try:
            In_Time.append((datetime.datetime.min + min(timings)).time())
        except:
            In_Time.append('ABSENT')
        try:
            Out_Time.append((datetime.datetime.min + max(timings)).time())
        except:
            Out_Time.append('ABSENT')
        try:
            actWrkHrs = calActualWorkingHours(ios, timings, doors, Level[len(Level)-1])
            if actWrkHrs == datetime.timedelta():
                Work_Duration.append('ABSENT')
            else:
                Work_Duration.append((datetime.datetime.min + actWrkHrs).time())
        except Exception as e:
            actWrkHrs = datetime.timedelta()
            Work_Duration.append('ABSENT')
        try:
            Total_Duration.append((datetime.datetime.min + max(timings)-min(timings)).time())
        except:
            Total_Duration.append('ABSENT')
        try:
            if actWrkHrs == absent:
                Non_Completed_Hours.append('ABSENT')
            elif actWrkHrs < StdWrkHrs :
                Non_Completed_Hours.append((datetime.datetime.min + StdWrkHrs-actWrkHrs).time())
            elif actWrkHrs >= StdWrkHrs :
                Non_Completed_Hours.append(datetime.time())
        except:
            if actWrkHrs == absent:
                Non_Completed_Hours.append('ABSENT')
            elif actWrkHrs >= StdWrkHrs :
                Non_Completed_Hours.append(datetime.time())
        if actWrkHrs >= StdWrkHrs :
            Remarks.append('COMPLETED')
            Additional_Hours.append((datetime.datetime.min + actWrkHrs-StdWrkHrs).time())
        elif actWrkHrs == absent:
            Remarks.append('ABSENT')
            Additional_Hours.append('ABSENT')
        elif actWrkHrs < StdWrkHrs :
            Remarks.append('NOT COMPLETED')
            Additional_Hours.append(datetime.time())
        del ios[:]; del timings[:]; del doors[:]

    #print(len(Artist_Code))
    #print(len(Artist_Name))
    #print(len(Department))
    #print(len(In_Time))
    #print(len(Out_Time))
    #print(len(Work_Duration))
    #print(len(Total_Duration))

    df = pd.DataFrame({'Artist Code':Artist_Code, 'Artist Name':Artist_Name, 'Department':Department, 'In Time':In_Time, 'Out Time':Out_Time, 'Work Duration':Work_Duration, 'Total Duration':Total_Duration, 'Non Completed Hours':Non_Completed_Hours, 'Additional Hours':Additional_Hours, 'Remarks':Remarks})

    from openpyxl import load_workbook
    #from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting import Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.styles import Alignment
    from openpyxl.styles.borders import Border, Side

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor='FFC7CE')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator='containsText', text='ABSENT', dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("ABSENT",E1)))']

    try:
        book = load_workbook('export_test.xlsx')
        writer = pd.ExcelWriter('export_test.xlsx', engine='openpyxl')
    except:
        writer = pd.ExcelWriter('export_test.xlsx', engine='openpyxl')
        book = writer.book
    #writer = pd.ExcelWriter('export_test.xlsx', engine='openpyxl')
    #workbook = writer.book
    writer.book = book
    df.to_excel(writer, sheet_name=formatDate(date))
    worksheet = writer.sheets[formatDate(date)]

    if worksheet in writer.book:
        del writer.book[worksheet]

    for row in worksheet['A1:K100']:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

    for cell in worksheet['A'] + worksheet[1]:
        cell.style = 'Pandas'

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    worksheet.conditional_formatting.add('E1:K500', rule)
    worksheet.sheet_properties.tabColor = "1072BA"

    writer.save()

def exportMonth(month, year):
    wd = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    StdWrkHrs = datetime.timedelta(hours=8, minutes=29, seconds=59)
    absent = datetime.timedelta()
    totalDays = calendar.monthrange(int(year), int(month))[1]
    #Month = calendar.monthcalendar(year, month)
    db = pymysql.connect("10.10.5.60", "mcheck", "mcheck@123", "essl", autocommit=True, connect_timeout=1)
    cur = db.cursor()
    cur1 = db.cursor()
    cur2 = db.cursor()

    artistID = []
    artistName = []
    artistDept = []
    artistLvl = []

    ios = []
    timings = []
    doors = []

    Work_Duration = []
    std_Hours = []
    Additional_Hours = []

    Artist_Leaves = []
    Month_Std_Hours = []
    Month_Completed_Hours = []
    Month_Non_Completed_Hours = []
    Month_Additional_Hours = []
    AH_Decimal = []
    AHD_Decimal = []

    cur.execute("SELECT ID, Name, Department, Level FROM essl.user_master WHERE ID != '1000' AND Status = 'OPEN'")

    for data in cur.fetchall():
        artistID.append(data[0])
        artistName.append(data[1])
        artistDept.append(data[2])
        artistLvl.append(data[3])

    for id in artistID:
        monthlyWrkHours.id.append(id)
        Artist_Leaves.append(str(monthlyWrkHours.calArtistLeaveMon(int(year), int(month))))

        monthlyPopup.month = (month+"-"+year)
        monthlyPopup.workTime(month+"-"+year)
        Month_Std_Hours.append(formatTime(monthlyPopup.workTime.tarWorkingTime))

        Month_Completed_Hours.append(formatTime(monthlyWrkHours.calMonWrkHrs(year, month)))

        if monthlyPopup.workTime.tarWorkingTime-monthlyWrkHours.calMonWrkHrs(year, month) < datetime.timedelta():
            Month_Non_Completed_Hours.append("00:00:00")
            Month_Additional_Hours.append(formatTime(monthlyWrkHours.calMonWrkHrs(year, month) - monthlyPopup.workTime.tarWorkingTime))
            AH_Decimal.append("%.2f"%(round((monthlyWrkHours.calMonWrkHrs(year, month) - monthlyPopup.workTime.tarWorkingTime).total_seconds()/3600, 2)))
            AHD_Decimal.append("%.2f"%(((monthlyWrkHours.calMonWrkHrs(year, month) - monthlyPopup.workTime.tarWorkingTime).total_seconds()) / 86399))
        else:
            Month_Non_Completed_Hours.append(formatTime(monthlyPopup.workTime.tarWorkingTime-monthlyWrkHours.calMonWrkHrs(year, month)))
            Month_Additional_Hours.append("00:00:00")
            AH_Decimal.append("0.0")
            AHD_Decimal.append("0.0")

    extras = pd.DataFrame({
        'Absents': Artist_Leaves,
        'AUG Actual Hours': Month_Std_Hours,
        'Completed Hours': Month_Completed_Hours,
        'Non Completed Hours': Month_Non_Completed_Hours,
        'Additional Completed Hours': Month_Additional_Hours,
        'Additional Hours Decimal': AH_Decimal,
        'Additional Hours in Days': AHD_Decimal,
        'OT Pay on Gross Pay(25%)': None,
        'Comments': None
    })

    """#topDict = {'INFORMATION': None}
    infoDict = {'Actual': None, 'Standard': None, 'Additional Hours': None}
    #for Week in Month:
    #    for Day in Week:
    #        topDict[Day]
    for Day in range(1, totalDays+1):
        Ddate = (str(year) + "-" + str(month) + "-" + str(Day))
        #Cdate = datetime.datetime.strptime(Ddate, '%Y-%m-%d')
        topDict[Ddate] = infoDict
    #print(topDict)
    df1 = pd.DataFrame(topDict)
    print(df1)
    df2 = pd.DataFrame(infoDict, index=[0])
    print(df2)"""

    from openpyxl import load_workbook
    from openpyxl.formatting import Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.styles import Alignment
    from openpyxl.styles.borders import Border, Side

    c = 4

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor='FFC7CE')
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator='containsText', text='ABSENT', dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("ABSENT",E1)))']

    try:
        book = load_workbook('export_test.xlsx')
        writer = pd.ExcelWriter('export_test.xlsx', engine='openpyxl')
    except:
        writer = pd.ExcelWriter('export_test.xlsx', engine='openpyxl')
        book = writer.book

    writer.book = book

    for Day in range(1, totalDays+1):
        Ddate = (str(year) + "-" + str(month) + "-" + str(Day))
        Cdate = datetime.datetime.strptime(Ddate, '%Y-%m-%d')

        for id, i in zip(artistID, range(len(artistID))):

            cur2.execute("SELECT IO, MTIME, DOOR FROM essl.`%d` WHERE MDate = '%s' ORDER BY MTIME ASC" %(int(id), Cdate))

            for dt in cur2.fetchall():
                ios.append(dt[0])
                timings.append(dt[1])
                doors.append(dt[2])

            try:
                print(artistLvl[i])
                #actWrkHrs = calActualWorkingHours(ios, timings, doors, artistLvl[len(artistLvl)-1])
                actWrkHrs = calActualWorkingHours(ios, timings, doors, artistLvl[i])
                if actWrkHrs == datetime.timedelta():
                    Work_Duration.append('00:00')
                else:
                    #print(actWrkHrs)
                    Work_Duration.append((datetime.datetime.min + actWrkHrs).time())
            except Exception as e:
                actWrkHrs = datetime.timedelta()
                Work_Duration.append('00:00')

            if actWrkHrs >= StdWrkHrs :
                Additional_Hours.append((datetime.datetime.min + actWrkHrs-StdWrkHrs).time())
            elif actWrkHrs == absent:
                Additional_Hours.append('00:00')
            elif actWrkHrs < StdWrkHrs :
                Additional_Hours.append('00:00')

            std_Hours.append(str(StdWrkHrs))

            del ios[:]; del timings[:]; del doors[:]

        dfs = pd.DataFrame({
            '%s'%(Ddate): Work_Duration,
            'Standard Hours': std_Hours,
            'Additional Hours': Additional_Hours
        })

        dfs.to_excel(writer, sheet_name=str(month), startrow=1, startcol=c, index=False)
        c += 3
        del Work_Duration[:]; del std_Hours[:]; del Additional_Hours[:]
        del dfs

    df = pd.DataFrame({
        'Artist Code': artistID,
        'Artist Name': artistName,
        'Department': artistDept
    })

    df.to_excel(writer, sheet_name=str(month), startrow=1)

    extras.to_excel(writer, sheet_name=str(month), startrow=1, startcol=c, index=False)
    worksheet = writer.sheets[str(month)]

    worksheet.freeze_panes= 'A2'
    worksheet.freeze_panes= 'E2'

    if worksheet in writer.book:
        del writer.book[worksheet]

    for row in worksheet['A1:DB100']:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

    for cell in worksheet['A'] + worksheet[1]:
        cell.style = 'Pandas'

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    worksheet.conditional_formatting.add('E1:K500', rule)
    worksheet.sheet_properties.tabColor = "1072BA"

    writer.save()
